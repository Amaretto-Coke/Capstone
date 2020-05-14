import pandas as pd
import os
from string import ascii_letters
import openpyxl
from openpyxl import load_workbook

def import_cases_and_fluids():
    # Creating a path string to the user interface excel file.
    path = os.path.dirname(os.getcwd()) + r'\UI.xlsm'

    # Importing the two Excel sheets as two new data frames.
    inputs = pd.read_excel(path, sheet_name='Parameters')

    # Setting the index to the property column.
    inputs.set_index('Property', drop=True, inplace=True)

    # Creating a dict that uses the index of the dataframe as keys.
    inputs = inputs.to_dict(orient='dict')['Value']

    # Returns a the inputs as a dictionary.
    return inputs


def export_results(dfs=None, df_names=None, open_after=False, index=False):
    path = os.path.dirname(os.getcwd()) + r'\Results.xlsx'
    
    if dfs is None:
        print('No date frames to export.')
    else:
        # Creating a path string to the user interface excel file
        if os.path.exists(path):
            pass
        else:
            wb = openpyxl.Workbook()
            wb.save(path)
            wb.close()
            
        for i in range(0, len(dfs)):
            sheet_name = validate_excel_sheet_name(df_names[i])
            writer = append_df_to_excel(filename=path,
                                        sheet_name=sheet_name,
                                        df=dfs[i],
                                        index=index)
            writer.save()
            writer.close()
        if open_after:
            os.system('start ' + path)


def append_df_to_excel(filename, df,
                       sheet_name='Sheet1',
                       start_row=None,
                       truncate_sheet=True,
                       **to_excel_kwargs):
    # Taken from https://stackoverflow.com/questions/38074678/append-existing-excel-sheet-with-new-dataframe-using-python-pandas
    """
    Append a DataFrame [df] to existing Excel file [filename]
    into [sheet_name] Sheet.
    If [filename] doesn't exist, then this function will create it.

    Parameters:
      filename : File path or existing ExcelWriter
                 (Example: '/path/to/file.xlsx')
      df : dataframe to save to workbook
      sheet_name : Name of sheet which will contain DataFrame.
                   (default: 'Sheet1')
      start_row : upper left cell row to dump data frame.
                 Per default (startrow=None) calculate the last row
                 in the existing DF and write to the next row...
      truncate_sheet : truncate (remove and recreate) [sheet_name]
                       before writing DataFrame to Excel file
      to_excel_kwargs : arguments which will be passed to `DataFrame.to_excel()`
                        [can be dictionary]

    Returns: None
    """

    # ignore [engine] parameter if it was passed
    if 'engine' in to_excel_kwargs:
        to_excel_kwargs.pop('engine')

    writer = pd.ExcelWriter(filename, engine='openpyxl')

    # Python 2.x: define [FileNotFoundError] exception if it doesn't exist
    try:
        FileNotFoundError
    except NameError:
        FileNotFoundError = IOError

    try:
        # try to open an existing workbook
        writer.book = load_workbook(filename)

        # get the last row in the existing Excel sheet
        # if it was not specified explicitly
        if start_row is None and sheet_name in writer.book.sheetnames:
            if truncate_sheet:
                start_row = 0
            else:
                start_row = writer.book[sheet_name].max_row

        # truncate sheet
        if truncate_sheet and sheet_name in writer.book.sheetnames:
            # index of [sheet_name] sheet
            idx = writer.book.sheetnames.index(sheet_name)
            # remove [sheet_name]
            writer.book.remove(writer.book.worksheets[idx])
            # create an empty sheet [sheet_name] using old index
            writer.book.create_sheet(sheet_name, idx)

        # copy existing sheets
        writer.sheets = {ws.title: ws for ws in writer.book.worksheets}
    except FileNotFoundError:
        # file does not exist yet, we will create it
        pass

    if start_row is None:
        start_row = 0

    # write out the new sheet
    df.to_excel(writer, sheet_name=sheet_name, startrow=start_row, **to_excel_kwargs)

    # save the workbook
    return writer
        
        
def validate_excel_sheet_name(val_string):
    allowed = ascii_letters + '1' + '2' + '3' + '4' + '5' + '6' + '7' + '8' + '9' + '0' + '-' + '_' + '.'
    if all(c in allowed for c in val_string):
        return val_string
    else:
        print('Excel sheet name characters must be alphanumeric or "-", "_", or a ".".')
        print('All not allowed character converted to "-".')
        for c in val_string:
            if c not in allowed:
                val_string = val_string.replace(c, '-')
        return val_string